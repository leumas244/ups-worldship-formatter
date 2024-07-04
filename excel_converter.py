from openpyxl import load_workbook
import re

from data_classes import Package

def get_packages_from_excel_file(excel_file: str) -> list[Package]:
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active

    titleRow = None
    senderColum = None
    reciverColum = None
    referenceColum = None
    packageCountColum = None

    for rowCounter in range(1,4):
        if not titleRow:
            for columCounter in range(1,7):
                cell = sheet.cell(row=rowCounter, column=columCounter)
                if cell.value == "Sender":
                    titleRow = rowCounter
                    senderColum = columCounter
                elif cell.value == "Empfänger":
                    titleRow = rowCounter
                    reciverColum = columCounter
                elif cell.value == "Variante / Farbe":
                    referenceColum = columCounter
                elif cell.value == "Menge":
                    packageCountColum = columCounter

    packages: list[Package] = []

    if titleRow and senderColum and reciverColum and referenceColum and packageCountColum:
        for row in range(titleRow+1, 200):
            if sheet.cell(row=row, column=reciverColum).value:
                reciverString = sheet.cell(row=row, column=reciverColum).value
                newPackage = Package(excelReciverString=reciverString, excel_row=row, excel_column=reciverColum)

                newPackage.referenceNumber = sheet.cell(row=row, column=referenceColum).value
                newPackage.packageCount = int(sheet.cell(row=row, column=packageCountColum).value)

                packages.append(newPackage)
            else:
                break
        
    return packages