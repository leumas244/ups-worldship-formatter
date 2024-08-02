import openpyxl
import warnings

import openpyxl.cell
import openpyxl.worksheet
import openpyxl.worksheet.merge
import openpyxl.worksheet.worksheet

from data_classes import Package
from address_parser import get_plz_city_and_region_from_line, get_highest_country_match
import settings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def get_workbook(excel_file_name: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(filename=excel_file_name)


def get_active_sheet_from_excel_file(
    excel_file_name: str,
) -> openpyxl.worksheet.worksheet.Worksheet:
    workbook = get_workbook(excel_file_name)
    return workbook.active


def cell_is_in_merge_cell_range(
    merged_cell_range: openpyxl.worksheet.merge.MergedCellRange,
    cell: openpyxl.cell.Cell,
) -> bool:
    min_row = merged_cell_range.min_row
    min_col = merged_cell_range.min_col
    max_row = merged_cell_range.max_row
    max_col = merged_cell_range.max_col

    if (min_row <= cell.row <= max_row) and (min_col <= cell.column <= max_col):
        return True
    else:
        return False


def is_cell_part_of_merged_cell(
    sheet: openpyxl.worksheet.worksheet.Worksheet, cell: openpyxl.cell.Cell
) -> bool:
    for merged_cell_range in sheet.merged_cells.ranges:
        if cell_is_in_merge_cell_range(merged_cell_range, cell):
            return True
    return False


def get_merged_cell_value(
    sheet: openpyxl.worksheet.worksheet.Worksheet, cell: openpyxl.cell.Cell
):
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        merged_cell_range: openpyxl.worksheet.merge.MergedCellRange
        for merged_cell_range in sheet.merged_cells.ranges:
            if cell_is_in_merge_cell_range(merged_cell_range, cell):
                start_cell: openpyxl.cell.Cell = merged_cell_range.start_cell
                return start_cell.value
    else:
        return cell.value


def are_cells_in_same_merged_cell(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    cell_1: openpyxl.cell.Cell,
    cell_2: openpyxl.cell.Cell,
) -> bool:
    if is_cell_part_of_merged_cell(sheet, cell_1) and is_cell_part_of_merged_cell(
        sheet, cell_2
    ):
        for merged_cell_range in sheet.merged_cells.ranges:
            if cell_is_in_merge_cell_range(merged_cell_range, cell_1):
                if cell_is_in_merge_cell_range(merged_cell_range, cell_2):
                    return True
        return False
    else:
        return False


def get_excel_type_from_reciverColum(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    titleRow: int,
    reciverColumn: int,
    trackingNumberColum: int,
):
    reciverCell = sheet.cell(row=titleRow, column=reciverColumn)
    reciverRightSideCell = sheet.cell(row=titleRow, column=(reciverColumn + 1))
    try:
        if are_cells_in_same_merged_cell(sheet, reciverCell, reciverRightSideCell):
            return "new_version"
        elif reciverColumn == 2 and trackingNumberColum == 6:
            return "old_version"
        else:
            return None
    except:
        return None


def all_necessary_information_is_available(sheet_informations: dict[str, int]) -> bool:
    for key, value in sheet_informations.items():
        if value is None:
            return False
    return True


def get_type_and_headerCells_from_excelSheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, int]:
    sheet_informations = {
        "excel_sheet_type": None,
        "titleRow": None,
        "senderColum": None,
        "reciverColum": None,
        "referenceColum": None,
        "packageCountColum": None,
        "shippingServiceColum": None,
        "trackingNumberColum": None,
    }

    for rowCounter in range(1, 5):
        if not sheet_informations["titleRow"]:
            for columCounter in range(1, 15):
                cell = sheet.cell(row=rowCounter, column=columCounter)
                if cell.value == "Sender":
                    sheet_informations["titleRow"] = rowCounter
                    sheet_informations["senderColum"] = columCounter
                elif cell.value == "Empfänger":
                    sheet_informations["titleRow"] = rowCounter
                    sheet_informations["reciverColum"] = columCounter
                elif cell.value == "Variante / Farbe":
                    sheet_informations["referenceColum"] = columCounter
                elif cell.value == "Menge":
                    sheet_informations["packageCountColum"] = columCounter
                elif cell.value == "Versand-Dienstleister":
                    sheet_informations["shippingServiceColum"] = columCounter
                elif cell.value == "Sendungs-Nummer":
                    sheet_informations["trackingNumberColum"] = columCounter
        else:
            break

    excel_type = get_excel_type_from_reciverColum(
        sheet,
        sheet_informations["titleRow"],
        sheet_informations["reciverColum"],
        sheet_informations["trackingNumberColum"],
    )
    sheet_informations["excel_sheet_type"] = excel_type

    if all_necessary_information_is_available(sheet_informations):
        return sheet_informations
    else:
        raise Exception(
            f"Die Excel Datei hat nicht das richtige Format um anylsiert zu werden"
        )


def get_packages_form_sheet_old_version(
    sheet: openpyxl.worksheet.worksheet.Worksheet, sheet_informations: dict[str, int]
) -> list[Package]:
    packages: list[Package] = []

    titleRow = sheet_informations["titleRow"]
    senderColum = sheet_informations["senderColum"]
    reciverColum = sheet_informations["reciverColum"]
    referenceColum = sheet_informations["referenceColum"]
    packageCountColum = sheet_informations["packageCountColum"]

    for row in range(titleRow + 1, 200):
        if sheet.cell(row=row, column=reciverColum).value:
            reciverString = sheet.cell(row=row, column=reciverColum).value
            coordinate = sheet.cell(row=row, column=reciverColum).coordinate
            newPackage = Package(
                excelReciverString=reciverString,
                excel_row=row,
                excel_column=reciverColum,
                coordinate=coordinate,
            )
            refrenceTuple: tuple = (
                sheet.cell(row=row, column=referenceColum).value,
                sheet.cell(row=row, column=packageCountColum).value,
            )
            newPackage.referenceNumbers.append(refrenceTuple)
            newPackage.packageCount = int(
                sheet.cell(row=row, column=packageCountColum).value
            )

            packages.append(newPackage)
        else:
            if sheet.cell(row=row, column=referenceColum).value:
                packages.remove(newPackage)
                refrenceTuple: tuple = (
                    sheet.cell(row=row, column=referenceColum).value,
                    sheet.cell(row=row, column=packageCountColum).value,
                )
                newPackage.referenceNumbers.append(refrenceTuple)
                newPackage.packageCount = newPackage.packageCount + (
                    int(sheet.cell(row=row, column=packageCountColum).value)
                )
                packages.append(newPackage)

            else:
                break

    return packages


def get_last_row_of_sender_cell(
    sheet: openpyxl.worksheet.worksheet.Worksheet, senderColumn: int, current_row: int
) -> int:
    cell = sheet.cell(row=current_row, column=senderColumn)
    for merged_cell_range in sheet.merged_cells.ranges:
        if cell_is_in_merge_cell_range(merged_cell_range, cell):
            return merged_cell_range.max_row
    raise Exception(
        f"Der Sender in Zeile {current_row} ist nicht mit Anderen Zellen verbunden"
    )


def there_is_package_information(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    senderColumn: int,
    reciverRightSideColumn: int,
    current_row: int,
) -> bool:
    last_row_of_sender_cell = get_last_row_of_sender_cell(
        sheet, senderColumn, current_row
    )
    value_counter = 0
    for row in range(current_row, (last_row_of_sender_cell + 1)):
        cell = sheet.cell(row=row, column=reciverRightSideColumn)
        if cell.value:
            value_counter += 1

    if value_counter >= 3:
        return True
    else:
        return False


def get_information_about_package_block(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    senderColumn: int,
    reciverRightSideColumn: int,
    current_row: int,
) -> dict[str]:
    result = {}
    result["last_row_of_sender_cell"] = get_last_row_of_sender_cell(
        sheet, senderColumn, current_row
    )
    result["there_is_package_information"] = there_is_package_information(
        sheet, senderColumn, reciverRightSideColumn, current_row
    )
    return result


def get_packages_form_sheet_new_version(
    sheet: openpyxl.worksheet.worksheet.Worksheet, sheet_informations: dict[str, int]
) -> list[Package]:
    package_has_an_error = [False, []]
    packages: list[Package] = []

    titleRow = sheet_informations["titleRow"]
    senderColum = sheet_informations["senderColum"]
    reciverColum = sheet_informations["reciverColum"]
    reciverRightSideColum = reciverColum + 1
    referenceColum = sheet_informations["referenceColum"]
    packageCountColum = sheet_informations["packageCountColum"]

    row = titleRow + 1
    while row < 200:
        block_info = get_information_about_package_block(
            sheet, senderColum, reciverRightSideColum, row
        )
        next_block_row = block_info["last_row_of_sender_cell"] + 1
        if block_info["there_is_package_information"]:
            is_there_error = False
            reciverString = "new_Excel_Version"
            coordinate = sheet.cell(row=row, column=reciverRightSideColum).coordinate
            newPackage = Package(
                excelReciverString=reciverString,
                excel_row=row,
                excel_column=reciverRightSideColum,
                coordinate=coordinate,
            )
            for block_row in range(row, next_block_row):
                try:
                    reciver_tag_cell = sheet.cell(row=block_row, column=reciverColum)
                    reciver_value_cell = sheet.cell(
                        row=block_row, column=reciverRightSideColum
                    )
                    reference_cell = sheet.cell(row=block_row, column=referenceColum)
                    packageCount_cell = sheet.cell(row=block_row, column=packageCountColum)

                    if reciver_tag_cell.value == "Name":
                        newPackage.recipientName = reciver_value_cell.value
                    elif reciver_tag_cell.value == "Firma":
                        newPackage.recipientNameAddtional = reciver_value_cell.value
                    elif reciver_tag_cell.value == "Adresse":
                        newPackage.address2 = reciver_value_cell.value
                    elif reciver_tag_cell.value == "(Adresse)":
                        newPackage.address1 = reciver_value_cell.value
                    elif reciver_tag_cell.value == "PLZ Ort":
                        region_info = get_plz_city_and_region_from_line(
                            reciver_value_cell.value
                        )
                        newPackage.state = region_info["state"]
                        newPackage.postalCode = region_info["postalCode"]
                        newPackage.city = region_info["city"]
                    elif reciver_tag_cell.value == "Land":
                        reciverValueStrip = reciver_value_cell.value.strip()
                        if reciver_tag_cell.value.strip() in settings.european_countrys:
                            newPackage.country = settings.european_countrys[
                                reciverValueStrip
                            ]
                        if not newPackage.country:
                            highest_country = get_highest_country_match(reciverValueStrip)
                            if highest_country[0] > 90:
                                newPackage.country = highest_country[1]
                            else:
                                raise Exception(
                                    f"Land '{reciverValueStrip}' wird nicht Unterstützt"
                                )
                    elif reciver_tag_cell.value == "Tel":
                        newPackage.phoneNumber = reciver_value_cell.value

                    if reference_cell.value:
                        refrenceTuple: tuple = (
                            reference_cell.value,
                            packageCount_cell.value,
                        )
                        newPackage.referenceNumbers.append(refrenceTuple)
                        if newPackage.packageCount:
                            newPackage.packageCount = newPackage.packageCount + packageCount_cell.value
                        else:
                            newPackage.packageCount = packageCount_cell.value
                    
                except Exception as e:
                    package_has_an_error[0] = True
                    package_has_an_error[1].append(newPackage)
                    is_there_error = True
                    print(f"Das Paket in Zeile {row}-{next_block_row-1} hat einen fehler: {str(e)}")
                    break
                    
            if not is_there_error:
                packages.append(newPackage)

            row = next_block_row
        else:
            break

    return (package_has_an_error, packages)


def get_packages_from_excel_file(excel_file_name: str) -> dict[str, list[Package]]:
    excel_file_has_a_problem = False
    sheet = get_active_sheet_from_excel_file(excel_file_name)

    sheet_info = get_type_and_headerCells_from_excelSheet(sheet)

    if sheet_info["excel_sheet_type"] == "new_version":
            excel_file_has_a_problem, packages = get_packages_form_sheet_new_version(sheet, sheet_info)

    elif sheet_info["excel_sheet_type"] == "old_version":
        packages = get_packages_form_sheet_old_version(sheet, sheet_info)

    result = {
        "type": sheet_info["excel_sheet_type"],
        "packages": packages,
        "excel_file_has_a_problem": excel_file_has_a_problem,
    }

    return result
