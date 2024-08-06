from datetime import datetime
import shutil
import xml.etree.ElementTree as ElementTree
import os
import re
import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet

import settings
from data_classes import Package
import excel_converter


def print_info(info: str, tab: int = 0, time_stamp: bool = True) -> None:
    columns, rows = shutil.get_terminal_size()
    columns -= 1
    pre_print_info = get_pre_print_info() + ("   " * tab)
    if time_stamp:
        print(pre_print_info, end="")
    else:
        print(" " * len(pre_print_info), end="")
    if (len(pre_print_info) + len(info)) > columns:
        result = split_string_by_length(info, (columns - len(pre_print_info)))
        for part in result:
            if result.index(part) == 0:
                print(part)
            else:
                print((" " * len(pre_print_info)) + part)
    else:
        print(info)


def get_pre_print_info() -> str:
    now_variable = datetime.now()
    pre_print_info = f"[{now_variable.strftime('%d/%m/%Y %H:%M:%S')}] "
    return pre_print_info


def split_string_by_length(s: str, length: int) -> list[str]:
    return [s[i : i + length] for i in range(0, len(s), length)]


def get_basename_from_file_path(path_to_file: str) -> str:
    return os.path.basename(path_to_file)


def get_all_processed_xml_files() -> list[str]:
    files = []
    out_pattern = r"\.Out$"
    try:
        for file in os.listdir(settings.xml_output_file_folder):
            if re.search(out_pattern, file):
                file_name = os.path.join(settings.xml_output_file_folder, file)
                if os.path.isfile(file_name):
                    files.append(file_name)

        return files

    except Exception as e:
        # print_info(f"Fehler beim Sammeln der zu analysierdenden Excel Dateien: {e}")
        # print_info(f"Das Programm beendet sich mit Problemen")
        exit()


def get_all_excel_folder() -> list[str]:
    folders = []
    try:
        for folder in os.listdir(settings.parsed_excel_file_folder):
            folder_name = os.path.join(settings.parsed_excel_file_folder, folder)
            if os.path.isdir(folder_name):
                folders.append(folder_name)
        return folders

    except Exception as e:
        # print_info(f"Fehler beim Sammeln der zu analysierdenden Excel Dateien: {e}")
        # print_info(f"Das Programm beendet sich mit Problemen")
        exit()


def get_all_excel_file_in_folder(excel_folder: str) -> str:
    found_excel_files = []
    xlsx_pattern = r"\.xlsx$"
    beginning_pattern = r"^~\$"
    for file in os.listdir(excel_folder):
        if re.search(xlsx_pattern, file):
            if re.search(beginning_pattern, file):
                return "file_is_used"

    for file in os.listdir(excel_folder):
        if re.search(xlsx_pattern, file):
            file_name = os.path.join(excel_folder, file)
            found_excel_files.append(file_name)

    if len(found_excel_files) == 1:
        return found_excel_files[0]
    else:
        return "no_file"


def get_matching_excel_and_out_files() -> list[tuple[str, str]]:
    matching_excel_and_out_files = []
    out_files = get_all_processed_xml_files()
    excel_folders = get_all_excel_folder()

    for excel_folder in excel_folders:
        folder_name = get_basename_from_file_path(excel_folder)
        for out_file_path in out_files:
            out_file_name = get_basename_from_file_path(out_file_path)
            if folder_name in out_file_name:
                excel_file_path = get_all_excel_file_in_folder(excel_folder)
                matching = (excel_file_path, out_file_path)
                matching_excel_and_out_files.append(matching)
                out_files.remove(out_file_path)
                break

    return matching_excel_and_out_files


def get_xml_tree(file_path: str) -> ElementTree.Element:
    tree = ElementTree.parse(file_path)
    return tree.getroot()


def create_trackingNumbers_and_refNumbers_assignment(package: Package) -> Package:

    if len(package.referenceNumbers) > 1:
        # ToDo Zeile manuell Prüfen lassen
        if len(package.referenceNumbers) == len(package.trackingNumbers):
            referenceNumbers = []
            for referenceNumber in package.referenceNumbers:
                newReferenceTuple = (referenceNumber[0], 1)
                referenceNumbers.append(newReferenceTuple)
            package.referenceNumbers = referenceNumbers

            trackingNumberCounter = 0
            for referenceNumber in package.referenceNumbers:
                package.excelTrackingAssignment["name"] = package.recipientName
                package.excelTrackingAssignment["referenceNumber"] = referenceNumber[0]
                package.excelTrackingAssignment["packageCount"] = referenceNumber[1]
                package.excelTrackingAssignment["trackingNumbers"] = [
                    package.trackingNumbers[trackingNumberCounter]
                ]
                trackingNumberCounter += 1
        else:
            # ToDo: Fehler melden, dass die Nummern Selbst eingetragen werden müssen
            pass

    elif len(package.referenceNumbers) == 1:
        package.excelTrackingAssignment["name"] = package.recipientName
        package.excelTrackingAssignment["referenceNumber"] = package.referenceNumbers[
            0
        ][0]
        package.excelTrackingAssignment["packageCount"] = package.referenceNumbers[0][1]
        package.excelTrackingAssignment["trackingNumbers"] = package.trackingNumbers
    else:
        # ToDo: Fehler keine Reference Nummer Gefunden
        pass

    return package


def get_element_index(parent: ElementTree.Element, target: ElementTree.Element) -> int:
    for idx, element in enumerate(parent):
        if element == target:
            return idx
    return -1


def get_rigth_xml_tree(out_file_path: str) -> ElementTree.Element:
    xxx_file_path = out_file_path.replace(".Out", ".xxx")

    out_tree = get_xml_tree(out_file_path)
    xxx_tree = get_xml_tree(xxx_file_path)

    for index, element in enumerate(out_tree):
        shipTo_out = element.find("{x-schema:OpenShipments.xdr}ShipTo")
        shipTo_xxx = xxx_tree[index].find("{x-schema:OpenShipments.xdr}ShipTo")
        shipTo_index = get_element_index(element, shipTo_out)
        element[shipTo_index] = shipTo_xxx

        shipFrom_out = element.find("{x-schema:OpenShipments.xdr}ShipFrom")
        shipFrom_xxx = xxx_tree[index].find("{x-schema:OpenShipments.xdr}ShipFrom")
        shipFrom_index = get_element_index(element, shipFrom_out)
        element[shipFrom_index] = shipFrom_xxx

        shipmentInformation_out = element.find(
            "{x-schema:OpenShipments.xdr}ShipmentInformation"
        )
        shipmentInformation_xxx = xxx_tree[index].find(
            "{x-schema:OpenShipments.xdr}ShipmentInformation"
        )
        shipmentInformation_index = get_element_index(element, shipmentInformation_out)
        element[shipmentInformation_index] = shipmentInformation_xxx

        package_out = element.find("{x-schema:OpenShipments.xdr}Package")
        package_xxx = xxx_tree[index].find("{x-schema:OpenShipments.xdr}Package")
        package_index = get_element_index(element, package_out)
        element[package_index] = package_xxx

    return out_tree


def get_proccesed_packages(xml_file_path: str) -> list[Package]:
    proccesed_packages: list[Package] = []
    xml_root = get_rigth_xml_tree(xml_file_path)

    openShipments = xml_root.findall("{x-schema:OpenShipments.xdr}OpenShipment")
    for openShipment in openShipments:
        process_status = openShipment.get("ProcessStatus")
        if process_status == "Processed":
            shipTo = openShipment.find("{x-schema:OpenShipments.xdr}ShipTo")
            shipmentInformation = openShipment.find(
                "{x-schema:OpenShipments.xdr}ShipmentInformation"
            )
            processMessage = openShipment.find(
                "{x-schema:OpenShipments.xdr}ProcessMessage"
            )
            package = Package(
                excelReciverString="", excel_row=0, excel_column=0, coordinate=""
            )
            package.recipientName = shipTo.find(
                "{x-schema:OpenShipments.xdr}CompanyOrName"
            ).text
            package.address1 = shipTo.find("{x-schema:OpenShipments.xdr}Address1").text
            package.postalCode = shipTo.find(
                "{x-schema:OpenShipments.xdr}PostalCode"
            ).text
            package.packageCount = int(
                shipmentInformation.find(
                    "{x-schema:OpenShipments.xdr}NumberOfPackages"
                ).text
            )

            for count in range(1, 6):
                ref_number_tree = processMessage.find(
                    ("{x-schema:OpenShipments.xdr}Reference" + str(count))
                )
                if ref_number_tree is not None:
                    if ref_number_tree.text:
                        refrence_tuple: tuple = (
                            ref_number_tree.text,
                            package.packageCount,
                        )
                        package.referenceNumbers.append(refrence_tuple)

            trackingNumbers = processMessage.find(
                "{x-schema:OpenShipments.xdr}TrackingNumbers"
            )
            for trackingNumber in trackingNumbers:
                package.trackingNumbers.append(trackingNumber.text)

            package = create_trackingNumbers_and_refNumbers_assignment(package)

            proccesed_packages.append(package)
    return proccesed_packages


def detect_packages_from_the_same_recipient(
    proccesed_packages: list[Package],
) -> list[Package]:
    detected_packages: list[Package] = []
    for proccesed_package in proccesed_packages:
        is_in_detected_packages = False
        for detected_package in detected_packages:
            if (
                proccesed_package.recipientName == detected_package.recipientName
                and proccesed_package.address1 == detected_package.address1
                and proccesed_package.postalCode == detected_package.postalCode
            ):
                if (
                    len(proccesed_package.referenceNumbers) == 1
                    and len(detected_package.referenceNumbers) == 1
                ):
                    if (
                        proccesed_package.referenceNumbers[0]
                        == detected_package.referenceNumbers[0]
                    ):
                        is_in_detected_packages = True
                        new_detected_package = detected_package
                        index_detected_package = detected_packages.index(
                            detected_package
                        )
                        new_detected_package.packageCount += (
                            proccesed_package.packageCount
                        )
                        new_refrence_tuple = (
                            detected_package.referenceNumbers[0][0],
                            new_detected_package.packageCount,
                        )
                        new_detected_package.referenceNumbers = [new_refrence_tuple]
                        for trackingNumber in proccesed_package.trackingNumbers:
                            new_detected_package.trackingNumbers.append(trackingNumber)

                        new_detected_package.excelTrackingAssignment["packageCount"] = (
                            new_detected_package.referenceNumbers[0][1]
                        )
                        new_detected_package.excelTrackingAssignment[
                            "trackingNumbers"
                        ] = new_detected_package.trackingNumbers

                        detected_packages[index_detected_package] = new_detected_package

                elif len(proccesed_package.referenceNumbers) > 1:
                    # ToDo: Fehler werfen
                    is_in_detected_packages = True
                    pass

        if not is_in_detected_packages:
            detected_packages.append(proccesed_package)

    return detected_packages


def get_merged_cell_value(sheet_, row, column):
    # Überprüfen Sie, ob die Zelle Teil einer zusammengeführten Zelle ist
    for merged_cell_range in sheet_.merged_cells.ranges:
        column_row_tuple = (row, column)
        if column_row_tuple in merged_cell_range.left:
            # Wenn ja, finden Sie die Zelle, die den Wert enthält
            return merged_cell_range.start_cell.value
    # Wenn die Zelle nicht Teil einer zusammengeführten Zelle ist, geben Sie den Zellenwert zurück
    return sheet_.cell(row=row, column=column).value


def store_ups_files_in_history(filepath: str) -> None:
    filepath_xxx = filepath.replace(".Out", ".xxx")
    try:
        file_name = os.path.basename(filepath)
        destination_file = os.path.join(settings.ups_hostory_folder, file_name)
        shutil.move(filepath, destination_file)

        file_name_xxx = os.path.basename(filepath_xxx)
        destination_file_xxx = os.path.join(settings.ups_hostory_folder, file_name_xxx)
        shutil.move(filepath_xxx, destination_file_xxx)
    except Exception as e:
        print_info(
            f"Konnte die Datei '{file_name}' nicht verschieben. Fehler: {str(e)}"
        )


def wirte_tracking_numbers_in_old_excel_version(
    workbook: openpyxl.Workbook,
    sheet_information: dict[str, int],
    formed_packages: list[Package],
) -> openpyxl.Workbook:
    sheet = workbook.active
    titleRow = sheet_information["titleRow"]
    reciverColum = sheet_information["reciverColum"]
    referenceColum = sheet_information["referenceColum"]
    packageCountColum = sheet_information["packageCountColum"]
    shippingServiceColum = sheet_information["shippingServiceColum"]
    trackingNumberColum = sheet_information["trackingNumberColum"]

    for row in range(titleRow + 1, 200):
        if sheet.cell(row=row, column=reciverColum).value:
            if not sheet.cell(row=row, column=trackingNumberColum).value:
                reciverString = sheet.cell(row=row, column=reciverColum).value
                packageCountValue = sheet.cell(row=row, column=packageCountColum).value
                refrencCoulmValue = sheet.cell(
                    row=row, column=referenceColum
                ).value.strip()

                for package in formed_packages:
                    if package.recipientName in reciverString:
                        if len(package.trackingNumbers) == 1:

                            if (
                                len(package.trackingNumbers) == packageCountValue
                                and package.referenceNumbers[0][0] == refrencCoulmValue
                            ):
                                sheet.cell(
                                    row=row, column=shippingServiceColum
                                ).value = "UPS"
                                sheet.cell(
                                    row=row, column=trackingNumberColum
                                ).value = package.trackingNumbers[0]
                                formed_packages.remove(package)
                                break
                        elif len(package.trackingNumbers) > 1:
                            if (
                                len(package.trackingNumbers) == packageCountValue
                                and package.referenceNumbers[0][0] == refrencCoulmValue
                            ):
                                sheet.cell(
                                    row=row, column=shippingServiceColum
                                ).value = "UPS"
                                out = ""
                                for trackingNumber in package.trackingNumbers:
                                    out = out + trackingNumber + ", "
                                sheet.cell(
                                    row=row, column=trackingNumberColum
                                ).value = out
                                formed_packages.remove(package)
                                break

        else:
            if sheet.cell(row=row, column=referenceColum).value:
                if not sheet.cell(row=row, column=trackingNumberColum).value:
                    reciverString = get_merged_cell_value(
                        sheet_=sheet, row=row, column=reciverColum
                    )
                    packageCountValue = sheet.cell(
                        row=row, column=packageCountColum
                    ).value
                    refrencCoulmValue = sheet.cell(
                        row=row, column=referenceColum
                    ).value.strip()

                    for package in formed_packages:
                        if package.recipientName in reciverString:
                            if len(package.trackingNumbers) == 1:
                                if (
                                    len(package.trackingNumbers) == packageCountValue
                                    and package.referenceNumbers[0][0]
                                    == refrencCoulmValue
                                ):
                                    sheet.cell(
                                        row=row, column=shippingServiceColum
                                    ).value = "UPS"
                                    sheet.cell(
                                        row=row, column=trackingNumberColum
                                    ).value = package.trackingNumbers[0]
                                    formed_packages.remove(package)
                                    break
                            elif len(package.trackingNumbers) > 1:
                                if (
                                    len(package.trackingNumbers) == packageCountValue
                                    and package.referenceNumbers[0][0]
                                    == refrencCoulmValue
                                ):
                                    sheet.cell(
                                        row=row, column=shippingServiceColum
                                    ).value = "UPS"
                                    out = ""
                                    for trackingNumber in package.trackingNumbers:
                                        out = out + trackingNumber + ", "
                                    sheet.cell(
                                        row=row, column=trackingNumberColum
                                    ).value = out
                                    formed_packages.remove(package)
                                    break

            else:
                break
    return workbook


def check_old_excel_list_on_trackingnumber_gaps(
    workbook: openpyxl.Workbook,
    sheet_information: dict[str, int],
) -> None:
    sheet = workbook.active
    titleRow = sheet_information["titleRow"]
    reciverColum = sheet_information["reciverColum"]
    referenceColum = sheet_information["referenceColum"]
    trackingNumberColum = sheet_information["trackingNumberColum"]

    for row in range(titleRow + 1, 200):
        if sheet.cell(row=row, column=reciverColum).value:
            if not sheet.cell(row=row, column=trackingNumberColum).value:
                print_info(f"- Bitte Zeile {row} selbst eintragen", tab=1)
        else:
            if sheet.cell(row=row, column=referenceColum).value:
                if not sheet.cell(row=row, column=trackingNumberColum).value:
                    print_info(f"- Bitte Zeile {row} selbst eintragen", tab=1)
            else:
                break
    return

def package_is_same_as_excel_block(
    package: Package,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    last_row_of_sender_cell: int,
    reciverColum: int,
    reciverRightSideColum: int,
    referenceColum: int,
    packageCountColum: int
) -> bool:
    same_Name = False
    same_company = False
    same_adress = False
    same_postalCode = False
    same_refnumbers = False
    
    refnumbers = []
    
    for block_row in range(row, last_row_of_sender_cell+1):
        reciver_tag_cell = sheet.cell(row=block_row, column=reciverColum)
        reciver_value_cell = sheet.cell(
            row=block_row, column=reciverRightSideColum
        )
        reference_cell = sheet.cell(row=block_row, column=referenceColum)
        packageCount_cell = sheet.cell(row=block_row, column=packageCountColum)

        if reciver_tag_cell.value == "Name":
            if reciver_value_cell.value and package.recipientName:
                if package.recipientName in reciver_value_cell.value:
                    same_Name = True
            else:
                same_Name = True
                            
        elif reciver_tag_cell.value == "Firma":
            if reciver_value_cell.value and package.recipientNameAddtional:
                if package.recipientNameAddtional in reciver_value_cell.value:
                    same_company = True
            else:
                same_company = True
                
        elif reciver_tag_cell.value == "(Adresse)":
            if package.address1 in reciver_value_cell.value:
                same_adress = True
            
        elif reciver_tag_cell.value == "PLZ Ort":
            if package.postalCode in reciver_value_cell.value:
                same_postalCode = True
                
        if reference_cell.value:
                    refrenceTuple: tuple = (
                        reference_cell.value,
                        packageCount_cell.value,
                    )
                    refnumbers.append(refrenceTuple)
                        
    if refnumbers == package.referenceNumbers:
        same_refnumbers = True
                    
    if same_Name and same_company and same_adress and same_postalCode and same_refnumbers:
        return True
    else:
        return False


def wirte_tracking_numbers_in_new_excel_version(
    workbook: openpyxl.Workbook, 
    sheet_informations: dict[str, int],
    formed_packages: list[Package],
) -> openpyxl.Workbook:
    sheet = workbook.active
    titleRow = sheet_informations["titleRow"]
    senderColum = sheet_informations["senderColum"]
    reciverColum = sheet_informations["reciverColum"]
    reciverRightSideColum = reciverColum + 1
    referenceColum = sheet_informations["referenceColum"]
    packageCountColum = sheet_informations["packageCountColum"]
    shippingServiceColum = sheet_informations["shippingServiceColum"]
    trackingNumberColum = sheet_informations["trackingNumberColum"]
    
    empty_package_counter = 0
    row = titleRow + 1
    while row < 1000:
        block_info = excel_converter.get_information_about_package_block(
            sheet, senderColum, reciverRightSideColum, row
        )
        last_row_of_sender_cell = block_info["last_row_of_sender_cell"]
        next_block_row = last_row_of_sender_cell + 1
        
        if block_info["there_is_package_information"]:
            empty_package_counter = 0
            for package in formed_packages:
                if package_is_same_as_excel_block(package, sheet, row, last_row_of_sender_cell, reciverColum, reciverRightSideColum, referenceColum, packageCountColum):
                    if not sheet.cell(row=row, column=trackingNumberColum).value:
                        sheet.cell(row=row, column=shippingServiceColum).value = "UPS"
                        out = ""
                        for trackingNumber in package.trackingNumbers:
                            if package.trackingNumbers.index(trackingNumber) < (len(package.trackingNumbers) - 1):
                                out = out + trackingNumber + ", "
                            else:
                                out = out + trackingNumber
                        sheet.cell(row=row, column=trackingNumberColum).value = out
                    formed_packages.remove(package)
                    break
                
            row = next_block_row
        else:
            if empty_package_counter < 3:
                empty_package_counter += 1
                row = next_block_row
            else:
                break
    return workbook


def check_new_excel_list_on_trackingnumber_gaps(
    workbook: openpyxl.Workbook, sheet_informations: dict[str, int]
) -> None:
    sheet = workbook.active
    titleRow = sheet_informations["titleRow"]
    senderColum = sheet_informations["senderColum"]
    reciverColum = sheet_informations["reciverColum"]
    reciverRightSideColum = reciverColum + 1
    trackingNumberColum = sheet_informations["trackingNumberColum"]

    empty_package_counter = 0
    row = titleRow + 1
    while row < 1000:
        block_info = excel_converter.get_information_about_package_block(
            sheet, senderColum, reciverRightSideColum, row
        )
        last_row_of_sender_cell = block_info["last_row_of_sender_cell"]
        next_block_row = last_row_of_sender_cell + 1
        
        if block_info["there_is_package_information"]:
            empty_package_counter = 0
            if not sheet.cell(row=row, column=trackingNumberColum).value:
                print_info(f"- Bitte Zeile {row} selbst eintragen", tab=1)

            row = next_block_row
        else:
            if empty_package_counter < 3:
                empty_package_counter += 1
                row = next_block_row
            else:
                break
    return


def start_routine() -> None:
    print_info(f"Starte das Importieren der Tracking-Nummern in Excel-Files")
    matching_excel_and_out_files = get_matching_excel_and_out_files()

    for excel_file_path, out_file_path in matching_excel_and_out_files:
        print_info(
            f"Bearbeite die Datei '{get_basename_from_file_path(excel_file_path)}'"
        )
        proccesed_packages = get_proccesed_packages(out_file_path)

        formed_packages = detect_packages_from_the_same_recipient(proccesed_packages)

        workbook = excel_converter.get_workbook(excel_file_path)
        sheet = workbook.active

        try:
            sheet_info = excel_converter.get_type_and_headerCells_from_excelSheet(sheet)
        except Exception as e:
            print_info(
                f"Die Trackingnummern für die Datei '{get_basename_from_file_path(excel_file_path)}' konnten nicht importiert werden. Fehler: '{str(e)}'", tab=1
            )
            continue

        if sheet_info["excel_sheet_type"] == "new_version":
            workbook = wirte_tracking_numbers_in_new_excel_version(workbook, sheet_info, formed_packages)
            check_new_excel_list_on_trackingnumber_gaps(workbook, sheet_info)

        elif sheet_info["excel_sheet_type"] == "old_version":
            workbook = wirte_tracking_numbers_in_old_excel_version(workbook, sheet_info, formed_packages)
            check_old_excel_list_on_trackingnumber_gaps(workbook, sheet_info)

        else:
            print_info(
                f"Die Trackingnummern für die Datei '{get_basename_from_file_path(excel_file_path)}' " /
                "konnten nicht importiert werden. Das Format der Excel-Liste wird nicht unterstützt.", tab=1
            )
            continue
        
        try:
            workbook.save(excel_file_path)
        except Exception as e:
            print_info(
                f"Die Datei '{get_basename_from_file_path(excel_file_path)}' " /
                "konnten nicht gespeichert werden. Ist die Datei vielleicht geöffnet? Fehler: '{str(e)}'", tab=1
            )
            continue
        
        store_ups_files_in_history(out_file_path)

    print_info(f"Beende das Importieren der Tracking-Nummern in Excel-Files")


start_routine()
